import os,random,array,json,getpass
from openpyxl import load_workbook, Workbook
import pandas as pd
from cryptography.fernet import Fernet
from hashlib import sha256
from termcolor import colored

def generateKey():
    if os.path.isfile('masterKey.key'):
        print(colored('\nYou already have a key',"red"))
        input(colored('Press a key to continue...\n',"dark_grey",attrs=['reverse', 'blink']))
    else:
        key = Fernet.generate_key()
        with open('masterKey.key', 'wb') as filekey:
            filekey.write(key)
def encryptFile():
    try:
        with open('masterKey.key', 'rb') as filekey:
            key = filekey.read()
        fernet = Fernet(key)
        with open('passwords.xlsx', 'rb') as file:
            original = file.read()
        encrypted = fernet.encrypt(original)
        with open('passwords.xlsx', 'wb') as encrypted_file:
            encrypted_file.write(encrypted)
    except:
        return
def decryptFile():
    try:
        with open('masterKey.key', 'rb') as filekey:
            key = filekey.read()
        fernet = Fernet(key)
        with open('passwords.xlsx', 'rb') as enc_file:
            encrypted = enc_file.read()
        decrypted = fernet.decrypt(encrypted)
        with open('passwords.xlsx', 'wb') as dec_file:
            dec_file.write(decrypted)
    except:
        return
def listPasswords():
    decryptFile()
    myFileName=r'passwords.xlsx'
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet']
    if ws.max_row == 1:
        print(colored("No password in the list\n","red"))
        return
    print(f"{pd.read_excel(myFileName)}\n\n")
    encryptFile()
    input(colored('Press a key to continue...\n',"dark_grey",attrs=['reverse', 'blink']))
def getInput():
    while True:
        try:
            website = input(colored("Enter website/name for the password i.e., 'google account': ","light_grey",attrs=['reverse', 'blink']))
            username= input(colored("Enter Username: ","light_grey",attrs=['reverse', 'blink']))
            password= input(colored("Enter Password: ","light_grey",attrs=['reverse', 'blink']))
            return [website,username,password]
        except:
            print(colored("wrong input try again","red"))
def addPassword():
    decryptFile()
    myFileName=r'passwords.xlsx'
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet']
    newRowLocation = ws.max_row +1
    myInput = getInput()
    ws.cell(column=1,row=newRowLocation, value=myInput[0])
    ws.cell(column=2,row=newRowLocation, value=myInput[1])
    ws.cell(column=3,row=newRowLocation, value=myInput[2])
    wb.save(filename=myFileName)
    wb.close()
    encryptFile()
def createXlsxFile():
    decryptFile()
    filepath = "passwords.xlsx"
    colnames = ['Website', 'Username', 'Password']
    wb = Workbook()
    wb.save(filepath)

    wb = load_workbook(filename=filepath)
    ws = wb['Sheet']
    ws.cell(column=1,row=1, value=colnames[0])
    ws.cell(column=2,row=1, value=colnames[1])
    ws.cell(column=3,row=1, value=colnames[2])
    wb.save(filename=filepath)
    wb.close()
    encryptFile()
def deletePassword():
    decryptFile()
    myFileName=r'passwords.xlsx'
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet']

    if ws.max_row == 1:
        print(colored("No password in the list to delete\n","red"))
        return
    try:
        row=int(input(colored("Enter the row number of the password you wish to delete: ","light_grey",attrs=['reverse', 'blink'])))+2
        if row <2 :
            print(colored("Row number can't be below 0\n","red"))
            return
        else:
            ws.delete_rows(row,1)
            wb.save(filename=myFileName)
            wb.close()
    except:
        print(colored("Must be a positive integer\n","green"))
    encryptFile()
def findPassword():
    decryptFile()
    myFileName=r'passwords.xlsx'
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet']
    if ws.max_row == 1:
        print(colored("No password in the list to find\n","red"))
        return

    passYouWant= input(colored("\n\nEnter the website, username, or password:","light_grey",attrs=['reverse', 'blink']))
    print(colored('\nWebsite : Username : Password',"light_grey",attrs=["reverse","blink"]))
    for row in ws.iter_cols(1):
        for cell in row:
            if cell.value == passYouWant:
                print(f"{ws.cell(row=cell.row, column=1).value} : {ws.cell(row=cell.row, column=2).value} : {ws.cell(row=cell.row, column=3).value}")
    encryptFile()
    input(colored('\nPress a key to continue...\n',"dark_grey",attrs=['reverse', 'blink']))
def generatePassword():
    decryptFile()
    try:
        website = input(colored("Enter website: ","light_grey",attrs=['reverse', 'blink']))
        username= input(colored("Enter Username: ","light_grey",attrs=['reverse', 'blink']))
        MAX_LEN = int(input(colored("Length of the password: ","light_grey",attrs=['reverse', 'blink'])))
        print("\n\n")
    except:
        print(colored("Invalid input\n","red"))

    DIGITS = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
    LOCASE_CHARACTERS = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h',
                        'i', 'j', 'k', 'm', 'n', 'o', 'p', 'q',
                        'r', 's', 't', 'u', 'v', 'w', 'x', 'y',
                        'z']
    UPCASE_CHARACTERS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H',
                        'I', 'J', 'K', 'M', 'N', 'O', 'P', 'Q',
                        'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
                        'Z']
    SYMBOLS = ['@', '#', '$', '%', '=', ':', '?', '.', '/', '|', '~', '>',
            '*', '(', ')', '<']
    COMBINED_LIST = DIGITS + UPCASE_CHARACTERS + LOCASE_CHARACTERS + SYMBOLS
    rand_digit = random.choice(DIGITS)
    rand_upper = random.choice(UPCASE_CHARACTERS)
    rand_lower = random.choice(LOCASE_CHARACTERS)
    rand_symbol = random.choice(SYMBOLS)

    temp_pass = rand_digit + rand_upper + rand_lower + rand_symbol
    for x in range(MAX_LEN - 4):
        temp_pass = temp_pass + random.choice(COMBINED_LIST)
        temp_pass_list = array.array('u', temp_pass)
        random.shuffle(temp_pass_list)
    password = ""
    for x in temp_pass_list:
            password = password + x
    myFileName=r'passwords.xlsx'
    wb = load_workbook(filename=myFileName)
    ws = wb['Sheet']
    newRowLocation = ws.max_row +1
    ws.cell(column=1,row=newRowLocation, value=website)
    ws.cell(column=2,row=newRowLocation, value=username)
    ws.cell(column=3,row=newRowLocation, value=password)
    wb.save(filename=myFileName)
    wb.close()
    encryptFile()
def setPassword():
    print(colored("\nWe will have to setup a master password. This password is unrecoverable","light_grey",attrs=["reverse","blink"]))
    master_password = getpass.getpass("Create a master password for the program: ")
    second_input = getpass.getpass("Verify your master pasword: ")
    if master_password == second_input:
        hash_master = sha256(master_password.encode("utf-8")).hexdigest()
        jfile = {"Master": {}}
        jfile["Master"] = hash_master
        with open("masterpassword.json", 'w') as jsondata:
            json.dump(jfile, jsondata, sort_keys=True, indent=4)
        print(colored("Password set successfully, ","green"))
    else:
        print(colored("Passwords do not match. Please try again","red"))
        return
def verifyPassword():
        if os.path.isfile("masterpassword.json"):
            with open("masterpassword.json", 'r') as masterPassword:
                jfile = json.load(masterPassword)

            stored_master_pass = jfile["Master"]
            try:
                master_password = getpass.getpass(colored("Enter Your Master Password: ","dark_grey",attrs=['reverse', 'blink']))
            except:
                print(colored("\nInvalid input","red"))
                return False

            if sha256(master_password.encode("utf-8")).hexdigest() == stored_master_pass:
                print(colored("Master password is correct",color="green"))
                return True
            else:
                print(colored("Master password is incorrect",color="red"))
                return False
def menu():
    try:
        choice = int(input(colored("\n1. List all passwords\n2. Add a password\n3. Delete a password\n4. Find a password\n5. Generate and add password\n6. Exit\n\ninput:","dark_grey",attrs=["reverse"])))
        print("\n\n")
        if 1<=choice<=6:
            if choice == 1:
                listPasswords()
            elif choice==2:
                addPassword()
            elif choice==3:
                deletePassword()
            elif choice==4:
                findPassword()
            elif choice == 5:
                generatePassword()
            elif choice==6:
                return 1
        else:
            print("Invalid number")
    except:
        print("Invalid choice")
def start():
    loggedIn=False
    while True:
        if os.path.isfile('passwords.xlsx') and os.path.isfile('masterkey.key') and os.path.isfile("masterpassword.json"):
            if loggedIn ==False:
                loggedIn=verifyPassword()
            else:
                start=menu()
                if start == 1:
                    print(colored("bye","light_grey",attrs=["reverse","blink"]))
                    break
        else:
            if os.path.isfile('passwords.xlsx')==False:
                createXlsxFile()
            if os.path.isfile('masterkey.key') == False:
                generateKey()
                encryptFile()
            if os.path.isfile("masterpassword.json")==False:
                setPassword()
if __name__ == "__main__":
    start()